from openpyxl import load_workbook
import regex as re

# CA excel Workbook and Worksheet used for input
CA_WORKBOOK = load_workbook('ca.xlsx')
CA_WORKSHEETS = CA_WORKBOOK.worksheets
# List sliced to start from 55 because main email data starts from Sheet 55
CA_WORKSHEETS = CA_WORKSHEETS[54:]

# Output Email Workbook And Worksheet
EMAILS_WORKBOOK = load_workbook('emails.xlsx')
EMAILS_WORKSHEET = EMAILS_WORKBOOK.active
# Value to be used in the print statement
VALUE = "Email"

# Regex used to find email from the cells
email_finder_regex = "(?<=E[-]?mail : ).*"
email_finder = re.compile(email_finder_regex)

# Starting read from a file so that if the program crashed for some reason, it can resume from where it left
with open('STARTING.txt', 'r') as R_File:
    EMAIL_CELL = "A"
    STARTING = R_File.read()


# Get All Emails from ca.xlsx
# return:
# all_emails_set - Set of all emails of ca.xlsx, Set is used to avoid duplicate emails
def get_all_emails() -> set:
    all_emails_set = set()
    for Worksheet in CA_WORKSHEETS:
        for Cells in Worksheet:
            for Cell in Cells:
                if Cell.value is not None:
                    Found_Emails = email_finder.findall(str(Cell.value))
                    if Found_Emails is not None and Found_Emails != []:
                        for Email in Found_Emails:
                            all_emails_set.add(Email)
    return all_emails_set


# params:
# Worksheet - Worksheet in which cell are to be added
# cell_name - Column Name in which values are to be added
# Starting - Row number from which to start filling in cells
# list_of_values - List of all the values which are to be added
def add_to_excel(Worksheet, cell_name, Starting, list_of_values):
    for value in list_of_values:
        cell = cell_name + Starting
        Worksheet[cell].value = value

        with open('STARTING.txt', 'w') as W_File:
            Starting = str(int(Starting) + 1)
            W_File.write(Starting)

        EMAILS_WORKBOOK.save('emails.xlsx')
        print(f"{VALUE}: {value} Added")


# Main Function
def main():
    list_of_all_emails = get_all_emails()
    add_to_excel(EMAILS_WORKSHEET, EMAIL_CELL, STARTING, list_of_all_emails)


if __name__ == "__main__":
    main()
