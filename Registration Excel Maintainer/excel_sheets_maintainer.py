"""
excel_sheets_maintainer
~~~~~~~~~~~~~~~~~~~~~~~

Maintains excel sheet of all form entries
Using email_content_parser, it gets all the entry details from emails of forms
and then updates them in the excel sheets
"""

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from email_content_parser import get_entry_details

# IMPORTANT Variables
filename = "Form Entries.xlsx"
last_n_days_entries = 2  # Get all form entries of last n days
gmail_label_name = "Form Entries"  # Label name in Gmail

# CONSTANTS
NAME_COLUMN = "A"
EMAIL_COLUMN = "B"
PHONE_COLUMN = "C"
CITY_COLUMN = "D"
DATE_COLUMN = "E"
ENTRY_FIELD_COLUMNS = (NAME_COLUMN, EMAIL_COLUMN, PHONE_COLUMN, CITY_COLUMN, DATE_COLUMN)

# Loading excel workbook
workbook = load_workbook(filename)
worksheet: Worksheet = workbook.worksheets[0]
empty_row_start = worksheet.max_row + 1


def get_worksheet_emails():
    emails: list = []
    for row_num in range(2, empty_row_start):
        email = worksheet[f"{EMAIL_COLUMN}{row_num}"].value
        emails.append(email)
    return emails


def update_excel_workbook():
    form_entries: list = get_entry_details(gmail_label_name, last_n_days_entries)
    last_updated_email = worksheet[f"{EMAIL_COLUMN}{empty_row_start - 1}"].value

    # Getting index from which entry to start adding in the worksheet
    # If the last updated entry is in the list, then start adding entries
    # from the entry after the last updated entry, identified by its email
    # If the last updated entry is not in the list, then add all the entries
    updation_starting_index = 0
    for index, (entry, date) in enumerate(form_entries):
        if entry[1] == last_updated_email:
            updation_starting_index = index + 1

    # Adding each entry to excel worksheet
    for entry_index, entry in enumerate(form_entries[updation_starting_index:]):
        excel_row = empty_row_start + entry_index
        for field_index in range(4):
            cell_coordinate = f"{ENTRY_FIELD_COLUMNS[field_index]}{excel_row}"
            worksheet[cell_coordinate] = entry[0][field_index]
        date_cell_coordinate = f"{DATE_COLUMN}{excel_row}"
        worksheet[date_cell_coordinate] = entry[1]

    workbook.save(filename=filename)


if __name__ == "__main__":
    update_excel_workbook()
